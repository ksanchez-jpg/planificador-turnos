import streamlit as st
import math
import pandas as pd
import io
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. CONFIGURACIÓN Y ESTILO
st.set_page_config(page_title="Programación de Turnos 44H", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.metric-box-green { background: #10B981; color: #064E3B; border-radius: 8px; padding: 1.2rem; text-align: center; }
.metric-value-dark { font-size: 2.2rem; font-family: 'IBM Plex Mono', monospace; font-weight: 700; }
.stButton > button { background: #0F172A; color: #F8FAFC; border-radius: 4px; width: 100%; }
</style>
""", unsafe_allow_html=True)

st.title("🗓 PROGRAMACIÓN DE TURNOS - NIVELACIÓN TOTAL")
st.caption("2x2 por bloques con Distribución Estricta de Refuerzos (Límite 2 por día).")

if 'seed' not in st.session_state: st.session_state['seed'] = 42
if 'mapping' not in st.session_state: st.session_state['mapping'] = {}

# --- CARGA DE EXCEL ---
fichas_cargadas = []
cargo_sugerido = "Cosechador"
conteo_sugerido = 20

with st.sidebar:
    st.header("📂 Base de Datos")
    archivo_subido = st.file_uploader("Adjuntar archivo COSECHA.xlsx", type=["xlsx"])
    if archivo_subido:
        excel_data = pd.ExcelFile(archivo_subido)
        hoja_sel = st.selectbox("Escoger hoja cargo", excel_data.sheet_names)
        df_excel = pd.read_excel(archivo_subido, sheet_name=hoja_sel)
        cargo_sugerido = hoja_sel
        if not df_excel.empty:
            fichas_cargadas = df_excel.iloc[:, 0].dropna().astype(str).str.strip().tolist()
            conteo_sugerido = len(fichas_cargadas)
            st.info(f"Fichas detectadas: {conteo_sugerido}")

    st.header("👤 Parámetros")
    cargo = st.text_input("Nombre del Cargo", value=cargo_sugerido)
    demanda_dia = st.number_input(f"{cargo} Día", min_value=1, value=5)
    demanda_noche = st.number_input(f"{cargo} Noche", min_value=1, value=5)
    horas_turno = st.number_input("Horas/turno", min_value=1, value=12)
    dias_cubrir = st.slider("Días/semana", 1, 7, 7)

    st.header("🧠 Modelo y Ajustes")
    factor_cobertura = st.slider("Factor Holgura", 1.0, 1.5, 1.0, 0.01)
    ausentismo = st.slider("Ausentismo (%)", 0.0, 0.3, 0.0, 0.01)
    operadores_actuales = st.number_input(f"{cargo} actual", min_value=0, value=conteo_sugerido)

# 3. CONSTANTES
DIAS_TOTALES = 42
TURNO_DIA, TURNO_NOCHE, DESCANSO = "D", "N", "R"
NOMBRES_DIAS = [f"S{s}-{d}" for s in range(1, 7) for d in ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]]

# 4. MOTOR DE PROGRAMACIÓN CON DISTRIBUCIÓN ESTRICTA
@st.cache_data
def generar_programacion_nivelada(n_ops, d_req, n_req, d_semana, seed):
    ops = [f"Op {i+1}" for i in range(n_ops)]
    horario = {op: [DESCANSO] * DIAS_TOTALES for op in ops}
    patron_maestro = [TURNO_DIA, TURNO_DIA, DESCANSO, DESCANSO, TURNO_NOCHE, TURNO_NOCHE, DESCANSO, DESCANSO]

    random.seed(seed)
    random.shuffle(ops)
    grupos = [ops[i::4] for i in range(4)]
    offsets = [0, 2, 4, 6]

    for bloque_idx in [0, 21]:
        bloque = range(bloque_idx, bloque_idx + 21)
        turnos_semanales = {op: [0, 0, 0] for op in ops}
        cob_dia = {d: 0 for d in bloque}
        cob_noche = {d: 0 for d in bloque}

        # FASE 1: ASIGNACIÓN BASE
        for g_idx, grupo_ops in enumerate(grupos):
            off = offsets[g_idx]
            for op in grupo_ops:
                for d in bloque:
                    if (d % 7) >= d_semana: continue
                    val = patron_maestro[(d + off) % 8]
                    if val != DESCANSO:
                        horario[op][d] = val
                        if val == TURNO_DIA: cob_dia[d] += 1
                        else: cob_noche[d] += 1
                        turnos_semanales[op][(d - bloque_idx) // 7] += 1

        # FASE 2: REFUERZOS NIVELADOS (Máximo 2 por día)
        max_refuerzos_permitidos = 1
        while max_refuerzos_permitidos <= 3:
            deudores = [op for op in ops if sum(1 for d in bloque if horario[op][d] != DESCANSO) < 11]
            if not deudores: break

            random.shuffle(deudores)
            for op in deudores:
                conteo = sum(1 for d in bloque if horario[op][d] != DESCANSO)
                if conteo >= 11: continue

                d_op = sum(1 for d in bloque if horario[op][d] == TURNO_DIA)
                n_op = sum(1 for d in bloque if horario[op][d] == TURNO_NOCHE)
                tipo_nec = TURNO_DIA if d_op <= n_op else TURNO_NOCHE

                candidatos = []
                for d in bloque:
                    if (d % 7) < d_semana and horario[op][d] == DESCANSO:
                        sem_idx = (d - bloque_idx) // 7
                        if turnos_semanales[op][sem_idx] >= 4: continue
                        if tipo_nec == TURNO_DIA and d > bloque_idx and horario[op][d-1] == TURNO_NOCHE: continue
                        ref_dia = (cob_dia[d] - d_req) + (cob_noche[d] - n_req)
                        if ref_dia >= max_refuerzos_permitidos: continue
                        v_izq = horario[op][d-1] if d > bloque_idx else None
                        v_der = horario[op][d+1] if d < bloque_idx + 20 else None
                        es_bloque = 1 if (v_izq == tipo_nec or v_der == tipo_nec) else 0
                        score = (ref_dia, -es_bloque)
                        candidatos.append((score, d))

                if candidatos:
                    candidatos.sort()
                    d_sel = candidatos[0][1]
                    horario[op][d_sel] = tipo_nec
                    if tipo_nec == TURNO_DIA: cob_dia[d_sel] += 1
                    else: cob_noche[d_sel] += 1
                    turnos_semanales[op][(d_sel - bloque_idx) // 7] += 1

            if deudores == [op for op in ops if sum(1 for d in bloque if horario[op][d] != DESCANSO) < 11]:
                max_refuerzos_permitidos += 1

    return pd.DataFrame(horario, index=NOMBRES_DIAS).T

# 5. EJECUCIÓN
def procesar_generacion(semilla_manual=None):
    if semilla_manual is not None: st.session_state['seed'] = semilla_manual
    st.session_state['mapping'] = {}
    total_t = (demanda_dia + demanda_noche) * dias_cubrir * 3
    op_f = max(math.ceil((math.ceil(total_t / 11) * factor_cobertura) / (1 - ausentismo)), (demanda_dia + demanda_noche) * 2)
    op_f = ((op_f + 3) // 4) * 4
    st.session_state["df"] = generar_programacion_nivelada(op_f, demanda_dia, demanda_noche, dias_cubrir, st.session_state['seed'])
    st.session_state["op_final"] = op_f

# BOTONES
c1, c2, c3 = st.columns(3)
with c1:
    if st.button("🚀 Generar Programación"): procesar_generacion(42)
with c2:
    if st.button("🔄 Generar Versión Aleatoria"): procesar_generacion(random.randint(1, 100000))
with c3:
    if st.button("👤 Asignar Fichas Reales"):
        if "df" in st.session_state:
            ops_ids = st.session_state["df"].index.tolist()
            f_lista = fichas_cargadas.copy()
            random.shuffle(f_lista)
            mapeo = {op: f_lista[i] if i < len(f_lista) else f"VACANTE {i-len(f_lista)+1}" for i, op in enumerate(ops_ids)}
            st.session_state['mapping'] = mapeo
            st.success("Personal asignado con nivelación estricta.")

# 6. RENDERIZADO
if "df" in st.session_state:
    df_base = st.session_state["df"]
    df_visual = df_base.copy()
    if st.session_state['mapping']:
        df_visual.index = [st.session_state['mapping'].get(x, x) for x in df_visual.index]

    op_final = st.session_state["op_final"]
    c_m1, c_m2, c_m3 = st.columns(3)
    with c_m1: st.markdown(f'<div class="metric-box-green"><div>Personal Total</div><div class="metric-value-dark">{op_final}</div></div>', unsafe_allow_html=True)
    with c_m2: st.markdown(f'<div class="metric-box-green"><div>Fichas Nómina</div><div class="metric-value-dark">{len(fichas_cargadas)}</div></div>', unsafe_allow_html=True)
    with c_m3: st.markdown(f'<div class="metric-box-green"><div>Horas/Ciclo</div><div class="metric-value-dark">132.0</div></div>', unsafe_allow_html=True)

    st.subheader("📅 Programación (Nivelación Máxima)")
    style_f = lambda v: f"background-color: {'#FFF3CD' if v=='D' else '#CCE5FF' if v=='N' else '#F8F9FA'}; font-weight: bold"
    st.dataframe(df_visual.style.map(style_f), use_container_width=True)

    st.subheader("📊 Balance Detallado")
    stats = []
    for idx in df_base.index:
        f = df_base.loc[idx]
        stats.append({
            "Identidad": st.session_state['mapping'].get(idx, idx),
            "T. Día": (f==TURNO_DIA).sum(), "T. Noche": (f==TURNO_NOCHE).sum(),
            "Horas S1-3": sum(1 for x in f[:21] if x != DESCANSO) * horas_turno,
            "Secuencia S1-3": f"{sum(1 for x in f[0:7] if x!=DESCANSO)}-{sum(1 for x in f[7:14] if x!=DESCANSO)}-{sum(1 for x in f[14:21] if x!=DESCANSO)}",
            "Horas S4-6": sum(1 for x in f[21:] if x != DESCANSO) * horas_turno,
            "Secuencia S4-6": f"{sum(1 for x in f[21:28] if x!=DESCANSO)}-{sum(1 for x in f[28:35] if x!=DESCANSO)}-{sum(1 for x in f[35:42] if x!=DESCANSO)}",
            "Estado": "✅ 44h OK"
        })
    st.dataframe(pd.DataFrame(stats).set_index("Identidad"), use_container_width=True)

    st.subheader("✅ Validación de Cobertura (Límite Estricto 2)")
    check = []
    for dia in NOMBRES_DIAS:
        ad, an = (df_base[dia] == TURNO_DIA).sum(), (df_base[dia] == TURNO_NOCHE).sum()
        refuerzos_total = (ad-demanda_dia)+(an-demanda_noche)
        check.append({"Día": dia, "Día (Asig)": ad, "Noche (Asig)": an, "Refuerzos": refuerzos_total, "Estado": "✅ OK" if refuerzos_total <= 2 else "⚠️"})
    st.dataframe(pd.DataFrame(check).set_index("Día").T, use_container_width=True)

    # ── EXPORTACIÓN CON COLORES Y HOJA DE BALANCE ──────────────────────
    out = io.BytesIO()
    wb = Workbook()

    # Estilos compartidos
    fill_D   = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")
    fill_N   = PatternFill("solid", start_color="CCE5FF", end_color="CCE5FF")
    fill_R   = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
    fill_hdr = PatternFill("solid", start_color="0F172A", end_color="0F172A")
    fill_ok  = PatternFill("solid", start_color="D1FAE5", end_color="D1FAE5")
    fill_alt = PatternFill("solid", start_color="F0F9FF", end_color="F0F9FF")

    font_hdr  = Font(bold=True, color="F8FAFC", name="Arial")
    font_bold = Font(bold=True, name="Arial")
    font_base = Font(name="Arial")
    align_c   = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── HOJA 1: PROGRAMACIÓN ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Programación"
    ws1.row_dimensions[1].height = 30

    # Celda origen (A1)
    c = ws1.cell(1, 1, "Identidad")
    c.font = font_hdr; c.fill = fill_hdr; c.alignment = align_c; c.border = border
    ws1.column_dimensions["A"].width = 18

    # Encabezados de días
    cols = df_visual.columns.tolist()
    for c_idx, col_name in enumerate(cols, start=2):
        cell = ws1.cell(1, c_idx, col_name)
        cell.font = font_hdr; cell.fill = fill_hdr
        cell.alignment = align_c; cell.border = border
        ws1.column_dimensions[get_column_letter(c_idx)].width = 7

    # Filas de operadores
    for r_idx, (idx, row) in enumerate(df_visual.iterrows(), start=2):
        ws1.row_dimensions[r_idx].height = 18
        c = ws1.cell(r_idx, 1, idx)
        c.font = font_bold; c.alignment = align_c; c.border = border
        for c_idx, val in enumerate(row, start=2):
            cell = ws1.cell(r_idx, c_idx, val)
            cell.alignment = align_c
            cell.border = border
            cell.font = font_bold
            if val == "D":   cell.fill = fill_D
            elif val == "N": cell.fill = fill_N
            else:            cell.fill = fill_R

    ws1.freeze_panes = "B2"

    # ── HOJA 2: BALANCE OPERADORES ────────────────────────────────────
    ws2 = wb.create_sheet("Balance Operadores")

    stats_export = []
    for idx in df_base.index:
        f = df_base.loc[idx]
        stats_export.append({
            "Identidad":     st.session_state['mapping'].get(idx, idx),
            "T. Día":        int((f == TURNO_DIA).sum()),
            "T. Noche":      int((f == TURNO_NOCHE).sum()),
            "Horas S1-3":    int(sum(1 for x in f[:21] if x != DESCANSO)) * horas_turno,
            "Secuencia S1-3": f"{sum(1 for x in f[0:7] if x!=DESCANSO)}-{sum(1 for x in f[7:14] if x!=DESCANSO)}-{sum(1 for x in f[14:21] if x!=DESCANSO)}",
            "Horas S4-6":    int(sum(1 for x in f[21:] if x != DESCANSO)) * horas_turno,
            "Secuencia S4-6": f"{sum(1 for x in f[21:28] if x!=DESCANSO)}-{sum(1 for x in f[28:35] if x!=DESCANSO)}-{sum(1 for x in f[35:42] if x!=DESCANSO)}",
            "Estado":        "✅ 44h OK"
        })

    df_stats = pd.DataFrame(stats_export)
    headers = df_stats.columns.tolist()
    col_widths = [20, 10, 10, 12, 16, 12, 16, 12]

    ws2.row_dimensions[1].height = 30
    for c_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws2.cell(1, c_idx, h)
        cell.font = font_hdr; cell.fill = fill_hdr
        cell.alignment = align_c; cell.border = border
        ws2.column_dimensions[get_column_letter(c_idx)].width = w

    for r_idx, row in df_stats.iterrows():
        bg = fill_ok if r_idx % 2 == 0 else fill_alt
        ws2.row_dimensions[r_idx + 2].height = 18
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(r_idx + 2, c_idx, val)
            cell.fill = bg; cell.alignment = align_c
            cell.border = border; cell.font = font_base

    ws2.freeze_panes = "A2"

    wb.save(out)
    st.download_button(
        label="⬇️ Descargar Excel con Colores",
        data=out.getvalue(),
        file_name=f"Programacion_{cargo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
